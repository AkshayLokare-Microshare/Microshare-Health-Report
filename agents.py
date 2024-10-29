#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

PATH = r'C:\Program Files (x86)\chromedriver.exe'
USERNAME = os.getenv("OPENSEARCH_EMAIL")
PASSWORD = os.getenv("OPENSEARCH_PASSWORD")

service = Service(PATH)
driver = webdriver.Chrome(service=service)

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Agents Dashboard is working':
                cell.fill = greenColor
            elif cell.value == 'Agents Dashboard is NOT working':
                cell.fill = redColor
            # elif cell.value == 'Error':
            #     cell.fill = blueColor

    workbook.save(excel_path)

def agents():

    with open(f'Health_Report({date}).txt', 'a') as file:
        driver.maximize_window()
        driver.get("https://prdlogs.microshare.io/app/login")

        wait = WebDriverWait(driver, 100)

        username_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-test-subj="user-name"]')))
        username_field.send_keys(USERNAME)

        password_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-test-subj="password"]')))
        password_field.send_keys(PASSWORD)

        yes_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'euiButton__text')))
        yes_button.click()

        time.sleep(10)
        form = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-test-subj="tenant-switch-radios"]')))

        global_button = wait.until(EC.presence_of_element_located((By.XPATH, "//label[@for='global']")))
        global_button.click()

        confirm_button = driver.find_element(By.XPATH, "//span[contains(text(),'Confirm')]").click()

        time.sleep(10)

        wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tbody')))
        # Wait for dashboard to load
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[data-test-subj="dashboardListingTitleLink-Dashboard---Stopped-Agents"]')))
        agent_dashboard = driver.find_element(By.CSS_SELECTOR, 'a[data-test-subj="dashboardListingTitleLink-Dashboard---Stopped-Agents"]')
        agent_dashboard.click()

        time.sleep(10)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "visualization")))

        # Refresh the page and wait for the element
        driver.refresh()
        time.sleep(10)  # Increased wait time to ensure page is fully refreshed

        # Get and parse the dashboard time element
        dashboard_time_str = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@role='gridcell']//span"))).text
        dashboard_time = datetime.fromisoformat(dashboard_time_str.replace('Z', '+00:00'))  # Parse the ISO 8601 date string

        current_date = datetime.now()

        # Calculate the difference in days
        date_diff = (current_date - dashboard_time).days

        file.write('\n\nAgents Dashboard:')

        # Print dates and times
        file.write(f'\nDashboard Date: {dashboard_time.strftime("%Y-%m-%d")} | Dashboard Time: {dashboard_time.strftime("%H:%M:%S")}\n')
        file.write(f'Current Date: {current_date.strftime("%Y-%m-%d")} | Current Time: {current_date.strftime("%H:%M:%S")}')

        if date_diff <= 7:
            file.write(f'\nAgents Dashboard is working - {date}')
            output = 'Agents Dashboard is working'
        else:
            file.write(f'\nAgents Dashboard is NOT working - {date}')
            output = 'Agents Dashboard is NOT working'

        time.sleep(2)
        driver.quit()

        append_output_to_excel(output, excel_output)
        color_output_in_excel(excel_output)

agents()
