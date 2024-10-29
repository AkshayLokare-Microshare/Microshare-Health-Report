#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

PATH = r'C:\Program Files (x86)\chromedriver.exe'
USERNAME = os.getenv("OPENSEARCH_EMAIL")
PASSWORD = os.getenv("OPENSEARCH_PASSWORD")

service = Service(PATH)
driver = webdriver.Chrome(service=service)

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Robot Dashboard is working':
                cell.fill = greenColor
            elif cell.value == 'Robot Dashboard is NOT working':
                cell.fill = redColor
            elif cell.value == 'Error':
                cell.fill = blueColor

    workbook.save(excel_path)

def robots():

    with open(f'Health_Report({date}).txt', 'a') as file:
        try:
            driver.maximize_window()
            driver.get("https://prdlogs.microshare.io/app/login")

            wait = WebDriverWait(driver, 100)

            username_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-test-subj="user-name"]')))
            username_field.send_keys(USERNAME)

            password_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-test-subj="password"]')))
            password_field.send_keys(PASSWORD)

            yes_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'euiButton__text')))
            yes_button.click()

            time.sleep(10)
            form = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-test-subj="tenant-switch-radios"]')))

            global_button = wait.until(EC.presence_of_element_located((By.XPATH, "//label[@for='global']")))
            global_button.click()

            confirm_button = driver.find_element(By.XPATH, "//span[contains(text(),'Confirm')]").click()
            
            time.sleep(10)

            wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tbody')))

            # Click on the target dashboard
            robot_dashboard = driver.find_element(By.CSS_SELECTOR, 'a[data-test-subj="dashboardListingTitleLink-Dashboard---Stopped-Robots"]').click()
            time.sleep(10)

            # Extract the timestamp
            dashboard_time_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-test-subj='dataGridRowCell'] span"))).text
            #print(f'Extracted Timestamp: {dashboard_time_element}')  

            current_time = datetime.now()
            file.write('\n\nRobots Dashboard:')

            # Define the timestamp format
            timestamp_format = "%Y-%m-%dT%H:%M:%S.%f"
            try:
                dashboard_datetime = datetime.strptime(dashboard_time_element, timestamp_format)
            except ValueError as e:
                print(f'Error parsing timestamp: {e}')
                print(f'Expected format: {timestamp_format}')
                print(f'Extracted timestamp: {dashboard_time_element}')
                driver.quit()
                exit()

            # Extract and format date and time components
            current_date = current_time.strftime("%Y-%m-%d")
            current_hour = current_time.strftime("%H")
            current_minute = current_time.strftime("%M")
            current_sec = current_time.strftime("%S")

            dashboard_date = dashboard_datetime.strftime("%Y-%m-%d")
            dashboard_hour = dashboard_datetime.strftime("%H")
            dashboard_minute = dashboard_datetime.strftime("%M")
            dashboard_sec = dashboard_datetime.strftime("%S")

            # Print results
            file.write(f'\nCurrent Date: {current_date} | Current Hour: {current_hour} | Current Min: {current_minute} | Current Sec: {current_sec}')
            file.write(f'\nLatest robot component Date: {dashboard_date} | Dashboard Hour: {dashboard_hour} | Dashboard Min: {dashboard_minute} | Dashboard Sec: {dashboard_sec}\n')

            # Calculate minute and second differences
            minute_diff = abs(int(current_minute) - int(dashboard_minute))
            second_diff = abs(int(current_sec) - int(dashboard_sec))

            # Check conditions and print appropriate messages
            if minute_diff <= 1 and second_diff <= 60:
                file.write(f'Robot Dashboard is working - {date}')
                output = 'Robot Dashboard is working'
                
            else:
                file.write(f'Robot Dashboard is NOT working - {date}')
                output = 'Robot Dashboard is NOT working'

        except ValueError as e:
            print(f'Error: {e}')
            output = 'Error'

        finally:
            driver.quit()

            append_output_to_excel(output, excel_output)
            color_output_in_excel(excel_output)

robots()
