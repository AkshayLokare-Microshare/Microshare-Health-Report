#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

PATH = r'C:\Program Files (x86)\chromedriver.exe'
USERNAME = os.getenv("OPENSEARCH_EMAIL")
PASSWORD = os.getenv("OPENSEARCH_PASSWORD")

service = Service(PATH)
driver = webdriver.Chrome(service=service)

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Applogs Dashboard is working':
                cell.fill = greenColor
            elif cell.value == 'Applogs Dashboard is NOT working':
                cell.fill = redColor
            elif cell.value == 'Error':
                cell.fill = blueColor

    workbook.save(excel_path)

def applogs():

    with open(f'Health_Report({date}).txt', 'a') as file:
        try:
            driver.maximize_window()
            driver.get("https://prdlogs.microshare.io/app/login")

            wait = WebDriverWait(driver, 100)

            username_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-test-subj="user-name"]')))
            username_field.send_keys(USERNAME)

            password_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-test-subj="password"]')))
            password_field.send_keys(PASSWORD)

            yes_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'euiButton__text')))
            yes_button.click()

            time.sleep(10)
            form = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-test-subj="tenant-switch-radios"]')))

            global_button = wait.until(EC.presence_of_element_located((By.XPATH, "//label[@for='global']")))
            global_button.click()

            confirm_button = driver.find_element(By.XPATH, "//span[contains(text(),'Confirm')]").click()
            
            time.sleep(10)

            wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tbody')))

            # Click on the target dashboard
            applogs_dashboard = driver.find_element(By.CSS_SELECTOR, 'a[data-test-subj="dashboardListingTitleLink-Dashboard---App-Logs"]').click()
            time.sleep(5)

            options_button = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Panel options for Discovery - App Logs']")))
            options_button.click()
            time.sleep(5)
            
            maximise = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Maximize panel')]")))
            maximise.click()
            time.sleep(5)

            timestamp = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@role='gridcell' and @data-test-subj='dataGridRowCell']//span/span")))
            
            # Example timestamp string: 2024-07-24T13:16:43.053
            timestamp_str = timestamp.text
            timestamp_now = datetime.now()

            # Parse the timestamp string into a datetime object
            timestamp_dt = datetime.strptime(timestamp_str, "%Y-%m-%dT%H:%M:%S.%f")

            # Format the datetime object into desired formats
            timestamp_date = timestamp_dt.strftime("%Y-%m-%d")
            timestamp_hour = timestamp_dt.strftime("%H")
            timestamp_min = timestamp_dt.strftime("%M")
            timestamp_sec = timestamp_dt.strftime("%S")

            current_date = timestamp_now.strftime("%Y-%m-%d")
            current_hour = timestamp_now.strftime("%H")
            current_minute = timestamp_now.strftime("%M")
            current_sec = timestamp_now.strftime("%S")

            # print("Date:", timestamp_date)        
            # print("Hours:", timestamp_hour)      
            # print("Minutes:", timestamp_min)  
            # print("Seconds:", timestamp_sec)  

            # print(f'\nApp Logs date: {timestamp_date}')
            # print(f'Current date: {current_date}')
            file.write(f'\n\nApplogs Dashboard:')
            file.write(f'\nLatest Log time: {timestamp_hour}:{timestamp_min}:{timestamp_sec}\n')
            file.write(f'Current time: {current_hour}:{current_minute}:{current_sec}')

            # Check if the current minute is more than 1 minute greater than the timestamp minute
            if int(current_minute) - int(timestamp_min) > 1 and int(current_sec) - int(timestamp_sec) > 60:
                print(f"\nAPP LOGS DASHBOARD needs a batch update")
                file.write(f'Applogs Dashboard is NOT working - {date}')
                output = 'Applogs Dashboard is NOT working'

            else:
                file.write(f'\nApplogs Dashboard is working - {date}')
                output = 'Applogs Dashboard is working'

        except ValueError as e:
            print(f'Error: {e}')
            output = 'Error'

        finally:
            time.sleep(2)
            driver.quit()

            append_output_to_excel(output, excel_output)
            color_output_in_excel(excel_output)

applogs()
