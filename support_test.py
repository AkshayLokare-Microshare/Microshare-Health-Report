#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

USERNAME = os.getenv("OUTLOOK_EMAIL")
PASSWORD = os.getenv("OUTLOOK_PASSWORD")
PATH = r"C:\Program Files (x86)\chromedriver.exe"

service = Service(PATH)
driver = webdriver.Chrome(service=service)

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Support Outlook is working':
                cell.fill = greenColor
            elif cell.value == 'Support Outlook is NOT working':
                cell.fill = redColor
            # elif cell.value == 'Error':
            #     cell.fill = blueColor

    workbook.save(excel_path)

def support_outlook():

    with open(f'Health_Report({date}).txt', 'a') as file:
        driver.get('https://outlook.office.com/mail/support@microshare.io/')

        wait = WebDriverWait(driver, 20)

        # Autofilling the username field
        username_field = wait.until(EC.presence_of_element_located((By.ID, 'i0116')))
        username_field.send_keys(USERNAME)
        username_field.send_keys(Keys.RETURN)

        time.sleep(2)

        # Autofilling the password field
        password_field = wait.until(EC.presence_of_element_located((By.ID, 'i0118')))
        password_field.send_keys(PASSWORD)
        ok_button = driver.find_element(By.ID, "idSIButton9").click()

        time.sleep(2)

        # Click the second "Yes" button to stay signed in
        ok_button_2 = wait.until(EC.presence_of_element_located((By.ID, "idSIButton9"))).click()

        # Wait for the mailbox to load
        time.sleep(5)  # Increase wait time if needed

        # Extract time texts
        time_texts = driver.find_elements(By.CSS_SELECTOR, 'span._rWRU')

        # Get current date and day
        current_datetime = datetime.now()
        current_day = current_datetime.strftime('%a')

        # This dict tells which day it was 2 days ago of the current day
        # The key is today's day and value is the day that was 2 days ago
        # Keep mondays value for 3 days
        day_map = {'Mon': 'Thu', 'Tue': 'Sun', 'Wed': 'Sat', 'Thu': 'Tue', 'Fri': 'Wed', 'Sat': 'Thu', 'Sun': 'Fri'}
        two_days_ago_day = day_map[current_day]

        print(f'Current Day: {current_day}')
        print(f'Day Two Days Ago: {two_days_ago_day}')
        file.write('\n\nSupport Outlook:')

        # Flag to check if any email is from two days ago
        email_found = False

        for time_element in time_texts:
            email_time_str = time_element.text
            email_day = email_time_str.split()[0]
            
            # Debugging print statements
            # print(f'Extracted Email Day: {email_day}')
            # print(f'Email Time String: {email_time_str}')
            
            if email_day == two_days_ago_day:
                email_found = True
                break  # No need to check further if we found an email from two days ago

        # Print result based on whether an email was found
        if not email_found:
            file.write(f'\nSupport Outlook is working - {date}')
            output = 'Support Outlook is working'

        else:
            file.write(f'\nSupport Outlook is NOT working - {date}')
            output = 'Support Outlook is NOT working'

        time.sleep(2)
        driver.quit()

        append_output_to_excel(output, excel_output)
        color_output_in_excel(excel_output)

support_outlook()
