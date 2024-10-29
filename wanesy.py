#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

# Path to ChromeDriver
PATH = r'C:\Program Files (x86)\chromedriver.exe'
USERNAME = os.getenv("WANESY_EMAIL")
PASSWORD = os.getenv("WANESY_PASSWORD")

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({time.strftime('%Y-%m-%d')}).xlsx'

def append_output_to_excel(output, excel_path):
    # Check if the file exists
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    #will print two lines after an unblank cell in column B
    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

    # Append the output
    # sheet.append([output])
    # workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Wanesy Login Success':
                cell.fill = greenColor
            elif cell.value == 'Wanesy Login Failed':
                cell.fill = redColor
            elif cell.value == 'Error':
                cell.fill = blueColor

    workbook.save(excel_path)

def wanesy_code():
    service = Service(PATH)
    driver = webdriver.Chrome(service=service)

    with open(f'Health_Report({date}).txt', 'a') as file:
        try:
            driver.get('https://microshare.wanesy.com')
            time.sleep(5)
            driver.refresh()

            wait = WebDriverWait(driver, 30)

            username = wait.until(EC.presence_of_element_located((By.ID, 'username')))
            username.send_keys(USERNAME)

            password = wait.until(EC.presence_of_element_located((By.ID, 'password')))
            password.send_keys(PASSWORD)

            login_button = driver.find_element(By.ID, 'loginButton')
            login_button.click()

            fleets = wait.until(EC.presence_of_element_located((By.XPATH, '//div[text() = "Gateways"]')))
            fleets.click()
            time.sleep(5)

            gateway_list = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tbody')))
            gateways = driver.find_elements(By.XPATH, '//tr/td[2]/a')

            if gateways:
                output = 'Wanesy Login Success'
                file.write(f'\n\nWanesy Login Success - {date}')
            else:
                output = 'Wanesy Login Failed'
                file.write(f'\n\nWanesy Login Failed - {date}')
            
        except Exception as e:
            output = 'Error'
            file.write(f'\nError: {e} - {date}')
            
        
        finally:
            time.sleep(2)
            driver.quit()

            # Append output to Excel and color it
            append_output_to_excel(output, excel_output)
            color_output_in_excel(excel_output)

wanesy_code()

