#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
import pyotp
from dotenv import load_dotenv
import os

PATH = r'C:\Program Files (x86)\chromedriver.exe'
service = Service(PATH)
driver = webdriver.Chrome(service=service)

USERNAME = os.getenv("OUTLOOK_EMAIL")
PASSWORD = os.getenv("HUBSPOT_PASSWORD")
SECRET_KEY = 'AGWWUGTHFGRMXW3XOOJNUDZGSXND6QR5'

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Hubspot is working':
                cell.fill = greenColor
            elif cell.value == 'Hubspot is NOT working':
                cell.fill = redColor
            elif cell.value == 'Error':
                cell.fill = blueColor

    workbook.save(excel_path)

def hubspot():
    with open(f'Health_Report({date}).txt', 'a') as file:
        try:
            driver.get('https://app.hubspot.com/contacts/437921/objects/0-5/views/all/list')
            driver.maximize_window()

            wait = WebDriverWait(driver, 30)

            username_field = wait.until(EC.presence_of_element_located((By.ID, 'username')))
            username_field.send_keys(USERNAME)

            password_field = wait.until(EC.presence_of_element_located((By.ID, 'password')))
            password_field.send_keys(PASSWORD)
            
            remember = driver.find_element(By.ID, 'remember')
            driver.execute_script("arguments[0].click();", remember)
            
            password_field.send_keys(Keys.RETURN)

            otp_field = wait.until(EC.presence_of_element_located((By.ID, 'code')))
            print("Please enter the OTP manually")
            
            # Get OTP from Google Authenticator
            totp = pyotp.TOTP(SECRET_KEY)
            otp = totp.now()
            # print(f'Your OTP is: {otp}')

            otp_field.send_keys(otp)
            otp_field.send_keys(Keys.RETURN)

            time.sleep(2)

            button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button[data-2fa-rememberme="true"]'))).click()

            time.sleep(10)

            grid_button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[data-test-id="page-type-board-button"]')))
            grid_button.click()
            time.sleep(5)

            pipeline_dropdown = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[data-selenium-test="pipelineSelector"]')))
            pipeline_dropdown.click()
            time.sleep(2)
            
            support_pipeline_option = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//span[@class='private-dropdown__item__label' and text()='Support Pipeline']"))
            )
            support_pipeline_option.click()
            time.sleep(2)

            waiting_on_contact_tickets = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-test-id='cdb-column-body'][data-rbd-droppable-id='4']")))
            time.sleep(5)
        
            if waiting_on_contact_tickets:
                title_elements = driver.find_elements(By.XPATH, '//a[@data-test-id="cdbc-title"]')

                if title_elements:
                    output = 'Hubspot is working'
                    file.write(f'\n\nHubspot is working - {date} \n')
                else:
                    output = 'Hubspot is NOT working'
                    file.write(f'\n\nHubspot is NOT working - {date} \n')

            new_tickets = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-test-id='cdb-column-body'][data-rbd-droppable-id='1']")))
            waiting_on_contact_tickets = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-test-id='cdb-column-body'][data-rbd-droppable-id='2']")))
            waiting_on_us_tickets = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-test-id='cdb-column-body'][data-rbd-droppable-id='3']")))

            # New bucket
            if new_tickets:
                for new_ticket in new_tickets:
                    new_ticket_names = new_ticket.find_elements(By.CSS_SELECTOR, 'a[data-test-id="cdbc-title"]')
                    created_at = new_ticket.find_elements(By.CSS_SELECTOR, 'span[data-test-id="cdbc-last-activity"]')

                    file.write(f'\nNew Bucket ({len(new_ticket_names)}):\n')
                    for new_ticket_name in new_ticket_names:
                        created_at_text = created_at[0].text if created_at else "N/A"
                        file.write(f'{new_ticket_name.text} | {created_at_text}\n')
            else:
                file.write('\n\nNew Bucket: No tickets found\n')

            # Waiting on Us bucket
            if waiting_on_us_tickets:
                for waiting_on_us_ticket in waiting_on_us_tickets:
                    waiting_on_us_ticket_names = waiting_on_us_ticket.find_elements(By.CSS_SELECTOR, 'a[data-test-id="cdbc-title"]')
                    created_at = waiting_on_us_ticket.find_elements(By.CSS_SELECTOR, 'span[data-test-id="cdbc-last-activity"]')

                    file.write(f'\nWaiting on Us ({len(waiting_on_us_ticket_names)}):\n')
                    for waiting_on_us_ticket_name in waiting_on_us_ticket_names:
                        created_at_text = created_at[0].text if created_at else "N/A"
                        file.write(f'{waiting_on_us_ticket_name.text} | {created_at_text}\n')
            else:
                file.write('\nWaiting on Us: No tickets found\n')

            # Waiting on Contact
            if waiting_on_contact_tickets:
                for waiting_on_contact_ticket in waiting_on_contact_tickets:
                    waiting_on_contact_ticket_names = waiting_on_contact_ticket.find_elements(By.CSS_SELECTOR, 'a[data-test-id="cdbc-title"]')
                    created_at = waiting_on_contact_ticket.find_elements(By.CSS_SELECTOR, 'span[data-test-id="cdbc-last-activity"]')

                    file.write(f'\nWaiting on Contact ({len(waiting_on_contact_ticket_names)}):\n')
                    for waiting_on_contact_ticket_name in waiting_on_contact_ticket_names:
                        created_at_text = created_at[0].text if created_at else "N/A"
                        file.write(f'{waiting_on_contact_ticket_name.text} | {created_at_text}\n')
            else:
                file.write('\nWaiting on Contact: No tickets found\n')


        except Exception as e:
            print(f'Error: {e} - {date}')
            file.write(f'Error: {e} - {date}')

        finally:
            time.sleep(2)
            driver.quit()

            append_output_to_excel(output, excel_output)
            color_output_in_excel(excel_output)
hubspot()
