#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

PATH = r'C:\Program Files (x86)\chromedriver.exe'
USERNAME = os.getenv("GRAFANA_EMAIL")
PASSWORD = os.getenv("GRAFANA_PASSWORD")

service = Service(PATH)
driver = webdriver.Chrome(service=service)

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Grafana Dashboard is working':
                cell.fill = greenColor
            elif cell.value == 'Grafana Dashboard is NOT working':
                cell.fill = redColor
            elif cell.value == 'Error':
                cell.fill = blueColor

    workbook.save(excel_path)

def grafana():
    def convert_to_float(value_str):
        value_str = value_str.replace(',', '').strip()  

        if 'K' in value_str:
            return float(value_str.replace('K', '').strip()) * 1000
        return float(value_str)

    with open(f'Health_Report({date}).txt', 'a') as file:
        try:
            driver.get("https://metrics.microshare.io/d/jwPKIsniz/strimzi-kafka-exporter?orgId=2&refresh=5s&viewPanel=12")

            wait = WebDriverWait(driver, 10)
            username_field = wait.until(EC.presence_of_element_located((By.NAME, 'user')))
            password_field = wait.until(EC.presence_of_element_located((By.NAME, 'password')))

            username_field.send_keys(USERNAME)
            password_field.send_keys(PASSWORD)
            password_field.send_keys(Keys.RETURN)

            # Wait for the table to load
            table = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, 'tbody'))
            )

            # Get the entire row (pod name, max value, current value)
            rows = table.find_elements(By.TAG_NAME, 'tr')

            file.write('\n\nGRAFANA DASHBOARD:\n')

            dashboard_working = False  # Flag to track if the dashboard is working

            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                
                if len(cells) >= 1:  
                    pod_name = cells[0].text.strip()  
                    max_value_str = cells[1].text.strip()
                    current_value_str = cells[2].text.strip()

                    if pod_name:
                        if not dashboard_working:  # Only write once
                            file.write(f'Grafana Dashboard is working - {date}')
                            output = 'Grafana Dashboard is working'
                            dashboard_working = True  # Set the flag to True
                        
                else:
                    print('Grafana is NOT working')
                    output = 'Grafana Dashboard is NOT working'

        except Exception as e:
            print(f"Error: {e}")
            output = 'Error'

        finally:
            time.sleep(2)
            driver.quit()

            append_output_to_excel(output, excel_output)
            color_output_in_excel(excel_output)


# Run the grafana function
grafana()

