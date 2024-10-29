from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

# Load environment variables
load_dotenv()

USERNAME = os.getenv("OUTLOOK_EMAIL")
PASSWORD = os.getenv("OUTLOOK_PASSWORD")
PATH = r"C:\Program Files (x86)\chromedriver.exe"

service = Service(PATH)
driver = webdriver.Chrome(service=service)

# Set the date format and Excel output file
date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    # Define colors
    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Support Outlook is working':
                cell.fill = greenColor
            elif cell.value == 'Support Outlook is NOT working':
                cell.fill = redColor

    workbook.save(excel_path)

def support_outlook():
    with open(f'Health_Report({date}).txt', 'a') as file:
        driver.get('https://outlook.office.com/mail/support@microshare.io/')
        wait = WebDriverWait(driver, 20)

        # Autofill the username and password fields
        username_field = wait.until(EC.presence_of_element_located((By.ID, 'i0116')))
        username_field.send_keys(USERNAME)
        username_field.send_keys(Keys.RETURN)

        time.sleep(2)

        password_field = wait.until(EC.presence_of_element_located((By.ID, 'i0118')))
        password_field.send_keys(PASSWORD)
        ok_button = driver.find_element(By.ID, "idSIButton9").click()

        time.sleep(2)

        ok_button_2 = wait.until(EC.presence_of_element_located((By.ID, "idSIButton9"))).click()

        # Wait for the mailbox to load
        time.sleep(5)

        # Extract time texts
        time_texts = driver.find_elements(By.CSS_SELECTOR, 'span._rWRU')

        # Get current date and time
        current_datetime = datetime.now()
        print(f"Current DateTime: {current_datetime}")  # Debug print

        file.write('\n\nSupport Outlook:')

        # Flag to check if any email is from two days ago
        email_found = False

        for time_element in time_texts:
            email_time_str = time_element.text
            print(f"Extracted Email Time String: {email_time_str}")  # Debug print

            # Check if time format is relative (e.g., "2h ago") or absolute (e.g., "Thu 3:00 PM")
            if 'ago' in email_time_str:
                # Extract the time difference in hours or minutes
                time_value, unit, _ = email_time_str.split()  # Example: '2 h ago'
                print(f"Relative Time: {time_value} {unit}")  # Debug print

                if unit == 'h':
                    hours_ago = int(time_value)
                    email_time = current_datetime - timedelta(hours=hours_ago)
                elif unit == 'm':
                    minutes_ago = int(time_value)
                    email_time = current_datetime - timedelta(minutes=minutes_ago)
                else:
                    email_time = current_datetime
            else:
                # Handle cases where the time is in a day and time format like "Thu 3:00 PM"
                try:
                    email_time = datetime.strptime(email_time_str, '%a %I:%M %p')
                    # The time is from an earlier day, adjust the date
                    email_time = email_time.replace(year=current_datetime.year)
                    if email_time > current_datetime:
                        email_time -= timedelta(days=7)  # Assuming email time is from the previous week
                except ValueError:
                    email_time = email_time_str  # Default to the text if parsing fails

            print(f"Parsed Email Time: {email_time}")  # Debug print

            # If the email time is within the last 24 hours
            if isinstance(email_time, datetime):
                time_diff = current_datetime - email_time
                if time_diff < timedelta(hours=24):
                    email_found = True
                    break  # No need to check further if we found a recent email

        # Print result based on whether an email was found
        if not email_found:
            file.write(f'\nSupport Outlook is working - {date}')
            output = 'Support Outlook is working'
        else:
            file.write(f'\nSupport Outlook is NOT working - {date}')
            output = 'Support Outlook is NOT working'

        time.sleep(2)
        driver.quit()

        append_output_to_excel(output, excel_output)
        color_output_in_excel(excel_output)

# Run the function
support_outlook()
