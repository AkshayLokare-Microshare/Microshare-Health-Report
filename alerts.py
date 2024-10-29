#okayyy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pytz
import time
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

load_dotenv()

# Credentials and path to chromedriver
USERNAME = os.getenv("OUTLOOK_EMAIL")
PASSWORD = os.getenv("OUTLOOK_PASSWORD")
PATH = r"C:\Program Files (x86)\chromedriver.exe"

print(f"username - {USERNAME} | password - {PASSWORD}")

date = time.strftime('%Y-%m-%d')
excel_output = f'health_status({date}).xlsx'

def append_output_to_excel(output, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet.title != 'Health Status':
        sheet.title = 'Health Status'

    row = sheet.max_row + 2
    sheet[f'B{row}'] = output
    workbook.save(excel_path)

def color_output_in_excel(excel_path):
    # Open workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    greenColor = PatternFill(start_color="0af790", end_color="0af790", fill_type="solid")
    redColor = PatternFill(start_color="ef0a0a", end_color="ef0a0a", fill_type="solid")
    blueColor = PatternFill(start_color="0adcef", end_color="0adcef", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Notifications are flowing in in Alerts Outlook':
                cell.fill = greenColor
            elif cell.value == 'Notifications are NOT flowing in in Alerts Outlook':
                cell.fill = redColor
            elif cell.value == 'Error':
                cell.fill = blueColor

    workbook.save(excel_path)

def alerts_outlook():
    # Setup ChromeDriver
    service = Service(PATH)
    driver = webdriver.Chrome(service=service)

    # Define ET timezone
    et_timezone = pytz.timezone('US/Eastern')

    driver.get('https://outlook.office365.com/mail/alert@microshare.io/AQMkADA5ZWUBN2NiLWJkMzgtNDVhNS1hYjgwLWM3MGE5MDg1MzRkYgAuAAADRItUqt8JkE6Y7QqIWeY2rAEAEdZkpzhNU06tw2g9UFyNswABIkcwzQAAAA%3D%3D/id/AAQkADA5ZWVlN2NiLWJkMzgtNDVhNS1hYjgwLWM3MGE5MDg1MzRkYgAQAPgj7zGHCFROpW5ACY3yfPM%3D')

    wait = WebDriverWait(driver, 20)

    username_field = wait.until(EC.presence_of_element_located((By.ID, 'i0116')))
    username_field.send_keys(USERNAME)
    username_field.send_keys(Keys.RETURN)

    time.sleep(2)

    password_field = wait.until(EC.presence_of_element_located((By.ID, 'i0118')))
    password_field.send_keys(PASSWORD)
    ok_button = driver.find_element(By.ID, "idSIButton9").click()

    time.sleep(2)

    ok_button_2 = wait.until(EC.presence_of_element_located((By.ID, "idSIButton9"))).click()

    # Get all emails 
    mail_list = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'MailList')))

    # Find email cards
    email_elements = driver.find_elements(By.CSS_SELECTOR, 'div.XG5Jd.TszOG')

    with open(f'Health_Report({date}).txt', 'a') as file:  # Open file in append mode
        file.write('\n\nALERTS OUTLOOK:\n')

        # Flag to check if notifications are flowing in
        notifications_flow_in = False

        for i in range(min(5, len(email_elements))):
            email = email_elements[i]
            
            # Extract the sender, subject, and time
            sender = email.find_element(By.CSS_SELECTOR, 'div.JBWmn.gy2aJ.CYQyC.Ejrkd span').get_attribute('title')
            subject = email.find_element(By.CSS_SELECTOR, 'div.IjzWp.XG5Jd.gy2aJ.Ejrkd.lME98 span').text
            time_received_str = email.find_element(By.CSS_SELECTOR, 'span._rWRU.Ejrkd.qq2gS.cbNn0').get_attribute('title')

            # Parse time_received_str into a datetime object with today's date
            try:
                # Parse the received time string
                time_received = datetime.strptime(time_received_str, '%I:%M %p')

                # Combine with today's date and set timezone to ET
                now = datetime.now(et_timezone)
                time_received = time_received.replace(year=now.year, month=now.month, day=now.day, tzinfo=et_timezone)

                # Calculate the time difference
                time_difference = now - time_received

                # Check if the time difference is greater than or equal to 4 hours
                if time_difference >= timedelta(hours=4):
                    file.write(f'Notifications are NOT flowing in in Alerts Outlook - {date}\n')
                    output = 'Notifications are NOT flowing in in Alerts Outlook'
                else:
                    if not notifications_flow_in:
                        file.write(f'Notifications are flowing in in Alerts Outlook - {date} (ET timezone)\n\n')
                        output = 'Notifications are flowing in in Alerts Outlook'
                        notifications_flow_in = True

                    file.write(f"Sender: {sender}\n")
                    file.write(f"Subject: {subject}\n")
                    file.write(f"Time: {time_received}\n")

            except ValueError as e:
                file.write(f'Error: {e} - {date}')
                output = 'Error'

        driver.quit()

    # Append output to Excel and color it
    append_output_to_excel(output, excel_output)
    color_output_in_excel(excel_output)

alerts_outlook()
